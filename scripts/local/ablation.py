
#25-out-4
# 0.1 0.4
# 0.2 0.4
# 0.3 0.4
# 0.4 0.4
# 0.5 0.4
# 0.6 0.4
# 0.7 0.4
# 0.8 0.4
# 0.9 0.4

# 1
# # 0.1 0.4 ~ 0.9 0.4
# means_25_out_4 = [0.6924577951, 0.6697208683, 0.7062017322, 0.6553830504, 0.6793076992
#                   0.6588614583, 0.6528378725, 0.6896581054, 0.6707389355]
# std_25_out_4 = [0.0138573784]


from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

wb = load_workbook(filename='./ablation.xlsx')
sheets = wb.sheetnames
ws = wb[sheets[0]]

data = np.zeros((9, 4))
for i in range(9):
    for j in range(4):
        data[i][j] = ws[chr(ord('A')+j) + str(i+1)].value
print(data)

deep_all_means = np.zeros((9,))+0.6429116925
deep_all_std = np.zeros((9,))+0.007890047566
rotation_means = np.zeros((9,))+0.6350216468
rotation_std = np.zeros((9,))+0.03174744516

#x = data[:,1]
x = [3, 4, 5, 25]
y = data[:,2]
yerr = data[:,3]
plt.errorbar(x=x, y=y, yerr=yerr, label='25-out-5')

plt.fill_between(x, deep_all_means-deep_all_std,
                 deep_all_means+deep_all_std,
                 facecolor='#fc9803', alpha=0.5,
                 edgecolor='none', label='deep all')
plt.plot(x, deep_all_means)

#plt.xlabel('β')
plt.xlabel('25-out-n')
plt.ylabel('Accuracy')
#plt.title('α=0.3')
plt.legend()
plt.grid(alpha=0.5)
plt.show()

